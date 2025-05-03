from app.funções.funções_conectar_email import aux_path_XML_destino, download_dir
from app.funções.funções_Gerais import obter_configs, navigate
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from openpyxl import Workbook
from datetime import datetime
from xml.dom import minidom
import pandas as pd
import flet as ft
import shutil
import os
import re

def extrair_MO(texto):
    padrao = r'MO:([A-Za-z0-9\-]+)'

    resultado = re.search(padrao, texto)

    if resultado:
        return resultado.group(1)
    else:
        return None

def formatar_planilha(worksheet):

    estilo_entrada = NamedStyle(name = 'estilo_entrada')
    estilo_entrada.font = Font(color = 'FFFFFF', name = 'consolas', size = 11)
    estilo_entrada.alignment = Alignment(horizontal = 'center', vertical = 'center')
    estilo_entrada.border = Border(left=Side(style = 'thin', color = 'FFFFFF'),
                                right=Side(style = 'thin', color = 'FFFFFF'),
                                top=Side(style = 'thin', color = 'FFFFFF'),
                                bottom=Side(style = 'thin', color = 'FFFFFF'))
    estilo_entrada.fill = PatternFill(start_color = '000000', end_color ='000000', fill_type = 'solid')
    for row in worksheet.iter_rows(min_row = 1, max_row = worksheet.max_row, min_col = 1, max_col = worksheet.max_column):
        for cell in row:  
            cell.style = estilo_entrada

def formatar_linha_difal(worksheet, cor_fundo, cor_fonte, cor_borda):
        
        for linha in worksheet.iter_rows(min_row = 1, max_row = 1, min_col = 1, max_col = 15):
            for celula in linha:
                celula.fill = PatternFill(start_color = cor_fundo, end_color = cor_fundo, fill_type = 'solid')
                celula.font = Font(color = cor_fonte, name = 'calibri', size = 10, bold = True)
                celula.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                celula.border = Border(left = Side (style = 'thin', color = cor_borda),
                                    right = Side (style = 'thin', color = cor_borda),
                                    top = Side (style = 'thin', color = cor_borda),
                                    bottom = Side (style = 'thin', color = cor_borda),
                                    )
                
# Botoes_entrada_Dell
def criar_planilha_entrada_nf_DELL(log_instance, page, views, current_view):

    navigate(current_view = current_view, page = page, views = views, view_name = 'logs e informações recebimento')

    botao_voltar = views['logs e informações recebimento'].content.controls[1].content.controls[0].controls[0]
    botao_voltar.disabled = True
    botao_voltar.content.color = ft.Colors.GREY_400
    page.update()

    log_instance.log_message('Iniciando criação da planilha Dell')

    if not os.path.exists(aux_path_XML_destino):
        log_instance.log_message('A pasta especificada não existe.')
    else:
        # Lista todos os arquivos na pasta
        arquivos = os.listdir(aux_path_XML_destino)

        # Cria uma nova planilha
        workbook = Workbook()
        sheet = workbook.active

        # Lista para armazenar os dados não formatados
        dados_nao_formatados = []

        # Itera sobre cada arquivo na pasta
        for idx, arquivo in enumerate(arquivos, start = 1):

            # Verifica se o caminho é um arquivo (e não um diretório)
            caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                        
                    # Abre o arquivo XML
                    with open(caminho_arquivo, 'r') as file:
                        nfe = minidom.parse(file)

                        # Código que extrai o número do chamado
                        chamado = nfe.getElementsByTagName('infCpl')
                        print_chamado = str((chamado[0].firstChild.data[30:39]))

                        # Código que extrai o número da NF
                        ref_nf = nfe.getElementsByTagName('nNF')
                        print_ref_nf = int((ref_nf[0].firstChild.data))

                        cProds = nfe.getElementsByTagName('cProd')
                        # Iterar sobre todos os elementos 'cProd' e adicionar à lista
                        produtos = [cProd.firstChild.nodeValue for cProd in cProds]

                        # Adiciona os dados à lista de dados não formatados
                        for produto in produtos:
                            dados_nao_formatados.append((print_chamado, print_ref_nf, produto))

                except AssertionError:
                    pass

                except Exception as e:
                    log_instance.log_message(f'Erro ao processar o arquivo {arquivo}: {e}')

        # Cria um DataFrame pandas com os dados
        df = pd.DataFrame(dados_nao_formatados, columns=['Chamado', 'NF', 'Produto'])

        # Ordena os números da coluna 'Chamado' e 'NF' do menor para o maior
        df['Chamado'] = pd.to_numeric(df['Chamado'])
        df['NF'] = pd.to_numeric(df['NF'])
        df = df.sort_values(['Chamado', 'NF'])

        # Remove as linhas duplicadas com base nos valores de 'Chamado', 'NF' e 'Produto'
        df_sem_duplicatas = df.drop_duplicates()

        # Converte o DataFrame de volta para uma lista de tuplas
        dados_sem_duplicatas = [tuple(x) for x in df_sem_duplicatas.values]

        # Insere os dados na planilha
        for dados in dados_sem_duplicatas:
            sheet.append(dados)

        # Mescla células com valores duplicados na coluna 'Chamado'
        row = 1
        while row <= sheet.max_row:
            chamado_atual = sheet.cell(row = row, column = 1).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column=1).value == chamado_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 1, end_row = row, end_column = 1)

            row += 1

        # Mescla células com valores duplicados na coluna 'NF'
        row = 1
        while row <= sheet.max_row:
            NF_atual = sheet.cell(row = row, column = 2).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column = 2).value == NF_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 2, end_row = row, end_column = 2)
                
            row += 1

        # Aplica formatação à planilha
        formatar_planilha(sheet)

        # Salva a planilha
        workbook.save('Planilha Notas de entrada_Dell.xlsx')

        if workbook.save:
            log_instance.log_message('Planilha criada com sucesso: Planilha Notas de entrada_Dell.xlsx')

            if os.path.exists(f'{download_dir}\\Planilha Notas de entrada_Dell.xlsx'):
                os.remove(f'{download_dir}\\Planilha Notas de entrada_Dell.xlsx')

            shutil.move('Planilha Notas de entrada_Dell.xlsx', download_dir)
        else:
            log_instance.log_message('Não foi possível criar a planilha')

        botao_voltar.disabled = False
        botao_voltar.content.color = ft.Colors.WHITE
        page.update()
            
# Botoes_entrada_HP
def criar_planilha_entrada_nf_HP(log_instance, page, views, current_view):

    navigate(current_view = current_view, page = page, views = views, view_name = 'logs e informações recebimento')

    botao_voltar = views['logs e informações recebimento'].content.controls[1].content.controls[0].controls[0]
    botao_voltar.disabled = True
    botao_voltar.content.color = ft.Colors.GREY_400
    page.update()

    log_instance.log_message('Iniciando criação da planilha HP')

    if not os.path.exists(aux_path_XML_destino):
        log_instance.log_message('A pasta especificada não existe.')
    else:
        # Lista todos os arquivos na pasta
        arquivos = os.listdir(aux_path_XML_destino)

        # Cria uma nova planilha
        workbook = Workbook()
        sheet = workbook.active

        # Lista para armazenar os dados não formatados
        dados_nao_formatados = []

        # Itera sobre cada arquivo na pasta
        for idx, arquivo in enumerate(arquivos, start = 1):

            # Verifica se o caminho é um arquivo (e não um diretório)
            caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)
            if os.path.isfile(caminho_arquivo):
                try:
                        
                    # Abre o arquivo XML
                    with open(caminho_arquivo, 'r') as file:
                        nfe = minidom.parse(file)

                        # Código que extrai o número do chamado
                        chamado = nfe.getElementsByTagName('infCpl')
                        if chamado:
                            print_chamado = chamado[0].firstChild.data
                            mo_valor = extrair_MO(print_chamado)
                        else:
                            mo_valor = 'MO não encontrado'

                        # Código que extrai o número da NF
                        ref_nf = nfe.getElementsByTagName('nNF')
                        print_ref_nf = int((ref_nf[0].firstChild.data))

                        cProds = nfe.getElementsByTagName('cProd')
                        # Iterar sobre todos os elementos 'cProd' e adicionar à lista
                        produtos = [cProd.firstChild.nodeValue for cProd in cProds]

                        # Adiciona os dados à lista de dados não formatados
                        for produto in produtos:
                            dados_nao_formatados.append((mo_valor, print_ref_nf, produto))

                except AssertionError:
                    pass

                except Exception as e:
                    log_instance.log_message(f'Erro ao processar o arquivo {arquivo}: {e}')

        # Cria um DataFrame pandas com os dados
        df = pd.DataFrame(dados_nao_formatados, columns=['Chamado', 'NF', 'Produto'])

        # Ordena os números da coluna 'Chamado' e 'NF' do menor para o maior
        df['Chamado'] = (df['Chamado'])
        df['NF'] = pd.to_numeric(df['NF'])
        df = df.sort_values(['Chamado', 'NF'])

        # Remove as linhas duplicadas com base nos valores de 'Chamado', 'NF' e 'Produto'
        df_sem_duplicatas = df.drop_duplicates()

        # Converte o DataFrame de volta para uma lista de tuplas
        dados_sem_duplicatas = [tuple(x) for x in df_sem_duplicatas.values]

        # Insere os dados na planilha
        for dados in dados_sem_duplicatas:
            sheet.append(dados)

        # Mescla células com valores duplicados na coluna 'Chamado'
        row = 1
        while row <= sheet.max_row:
            chamado_atual = sheet.cell(row = row, column = 1).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column=1).value == chamado_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 1, end_row = row, end_column = 1)

            row += 1

        # Mescla células com valores duplicados na coluna 'NF'
        row = 1
        while row <= sheet.max_row:
            NF_atual = sheet.cell(row = row, column = 2).value
            row_inicio = row

            while row < sheet.max_row and sheet.cell(row = row + 1, column = 2).value == NF_atual:
                row += 1

            if row - row_inicio > 0:
                sheet.merge_cells(start_row = row_inicio, start_column = 2, end_row = row, end_column = 2)
                
            row += 1

        # Aplica formatação à planilha
        formatar_planilha(sheet)

        # Salva a planilha
        workbook.save('Planilha Notas de entrada_HP.xlsx')

        if workbook.save:

            log_instance.log_message('Planilha criada com sucesso: Planilha Notas de entrada_HP')

            if os.path.exists(f'{download_dir}\\Planilha Notas de entrada_HP.xlsx'):
                os.remove(f'{download_dir}\\Planilha Notas de entrada_HP.xlsx')

            shutil.move('Planilha Notas de entrada_HP.xlsx', download_dir)

        else:
            log_instance.log_message('Não foi possível criar a planilha')

        botao_voltar.disabled = False
        botao_voltar.content.color = ft.Colors.WHITE
        page.update()

# Botoes_difal
def criar_planilha_difal(log_instance, page, views, current_view):

    navigate(current_view = current_view, page = page, views = views, view_name = 'logs e informações difal')

    botao_voltar = views['logs e informações difal'].content.controls[1].content.controls[0].controls[0]
    botao_voltar.disabled = True
    botao_voltar.content.color = ft.Colors.GREY_400
    page.update()

    aliquota_interna = obter_configs().get('aliquota_interna')
    nome_credenciada = obter_configs().get('nome_credenciada')
    cnpj_credenciada = obter_configs().get('cnpj_credenciada')

    log_instance.log_message('Iniciando criação da planilha Difal')

    if not os.path.exists(aux_path_XML_destino):
        log_instance.log_message('A pasta especificada não existe.')
    else:
        # Lista todos os arquivos na pasta
        arquivos = os.listdir(aux_path_XML_destino)

        #Cria um novo objeto
        workbook = Workbook()

        sheet = workbook.active

        #Adiciona dados à planilha
        Dados_linha1 = [
        'Nome Credenciada / Destino',
        'CNPJ Credenciada / Destino',
        'Mês de Referência',
        'Número da ND',
        'CNPJ Origem',
        'Razão Social / Origem',
        'Nr. Chamado do Cliente',
        'Número da NF de referência',
        'Data de Emissão',
        'Valor da BC do ICMS da NF de referência',
        'Alíquota Interestadual',
        'Valor do ICMS da NF de referência',
        'Alíquota interna do Estado de destino',
        'Alíquota do Diferencial',
        'Valor do ICMS recolhido por Diferencial de Alíquota'
        ]
        for col, valor in enumerate(Dados_linha1, start = 1):
            sheet.cell(row = 1, column = col, value = valor)

            #Adiciona a largura e altura das colunas
            largura_colunas = [
            25.43, 18.71, 10.00, 9.86, 20.14, 29.00, 18.43, 14.00, 12.14, 23.71, 16.86, 20.57, 15.00, 15.00, 22.57
            ]

        for i, largura in enumerate(largura_colunas, start = 1):
            sheet.column_dimensions[chr(64 + i)].width = largura
                                
        for a in range(2):
            sheet.row_dimensions[a].height = 34.00

        formatar_linha_difal(sheet, cor_fundo = '9BC2E6', cor_fonte = '000000', cor_borda = '000000')

        for idx, arquivo in enumerate(arquivos, start = 2):

            if arquivo.endswith(('procNFe.xml', 'nfe.xml')):

                try:
                    # Verifica se o caminho é um arquivo (e não um diretório)
                    caminho_arquivo = os.path.join(aux_path_XML_destino, arquivo)
                    if os.path.isfile(caminho_arquivo):

                        # Abre o arquivo XML
                        with open(caminho_arquivo, 'r') as file:
                            nfe = minidom.parse(file)

                            if arquivo.endswith('procNFe.xml'):
                                # Código que extrai o número do chamado
                                chamado = nfe.getElementsByTagName('infCpl')
                                print_chamado = str((chamado[0].firstChild.data[30:39]))
                                                        
                                # Código que extrai a data de emissão
                                data_emissao = nfe.getElementsByTagName('dhEmi')
                                print_data_emissao = str((data_emissao[0].firstChild.data[2:10]))
                                data = datetime.strptime(print_data_emissao, '%y-%m-%d')
                                nova_data = data.strftime('%d/%m/%y')
                                                        
                                # Código que extrai o ICMS da nota
                                ICMS_prod = nfe.getElementsByTagName('pICMS')
                                print_ICMS_prod = float(ICMS_prod[0].firstChild.data) / 100

                                # Código que extrai a nota fiscal da nota
                                ref_nf = nfe.getElementsByTagName('nNF')
                                print_ref_nf = str(ref_nf[0].firstChild.data)

                                # Código que extrai o valor da nota
                                vNF = nfe.getElementsByTagName('vNF')
                                print_valor_vNF = float(vNF[0].firstChild.data)

                                # Código para formatar o CNPJ
                                CNPJ_dell = nfe.getElementsByTagName('CNPJ')
                                print_CNPJ_dell = str(CNPJ_dell[0].firstChild.data)

                                print_CNPJ_dell_formatado = f"{print_CNPJ_dell[:2]}.{print_CNPJ_dell[2:5]}.{print_CNPJ_dell[5:8]}/{print_CNPJ_dell[8:12]}-{print_CNPJ_dell[12:]}"

                                # Código para formatar a razão social
                                razao_social_dell = nfe.getElementsByTagName('xNome')
                                print_razao_social_dell = str(razao_social_dell[0].firstChild.data)

                            elif arquivo.endswith('nfe.xml'):
                                try:
                                    chamado = nfe.getElementsByTagName('infCpl')
                                    if chamado:
                                        print_chamado = chamado[0].firstChild.data
                                        mo_valor = extrair_MO(print_chamado)

                                    else:
                                        mo_valor = 'MO não encontrado'
                                
                                except AssertionError:
                                    pass
                    
                                except Exception as e:
                                    log_instance.log_message(f'MO não contrado no arquivo: {arquivo}')
                                    continue

                                # Código que extrai a data de emissão
                                data_emissao = nfe.getElementsByTagName('dhEmi')
                                print_data_emissao = str((data_emissao[0].firstChild.data[2:10]))
                                data = datetime.strptime(print_data_emissao, '%y-%m-%d')
                                nova_data = data.strftime('%d/%m/%y')
                                            
                                # Código que extrai o ICMS da nota
                                ICMS_prod = nfe.getElementsByTagName('pICMS')
                                print_ICMS_prod = float(ICMS_prod[0].firstChild.data) / 100

                                # Código que extrai a nota fiscal da nota
                                ref_nf = nfe.getElementsByTagName('nNF')
                                print_ref_nf = str(ref_nf[0].firstChild.data)

                                # Código que extrai o valor da nota
                                vNF = nfe.getElementsByTagName('vNF')
                                print_valor_vNF = float(vNF[0].firstChild.data)
                                            
                                # Código para formatar o CNPJ
                                CNPJ_HP = nfe.getElementsByTagName('CNPJ')              
                                print_CNPJ_HP = str(CNPJ_HP[0].firstChild.data)

                                print_CNPJ_HP_formatado = f"{print_CNPJ_HP[:2]}.{print_CNPJ_HP[2:5]}.{print_CNPJ_HP[5:8]}/{print_CNPJ_HP[8:12]}-{print_CNPJ_HP[12:]}"

                                # Código para formatar a razão social
                                razao_social_hp = nfe.getElementsByTagName('xFant')
                                print_razao_social_hp = str(razao_social_hp[0].firstChild.data)
                                
                            # Código para formatar a coluna M da alíquota interna do estado (altera de estado para estado)
                            print_aliquota_interna = float(aliquota_interna) / 100

                            # Insere os dados na planilha
                            sheet[f'A{idx}'] = nome_credenciada
                            sheet[f'B{idx}'] = cnpj_credenciada
                                                        
                            if arquivo.endswith('procNFe.xml'):
                                sheet[f'E{idx}'] = print_CNPJ_dell_formatado
                                sheet[f'F{idx}'] = print_razao_social_dell
                                sheet[f'G{idx}'] = print_chamado 
                                
                            elif arquivo.endswith('-nfe.xml'):
                                sheet[f'E{idx}'] = print_CNPJ_HP_formatado
                                sheet[f'F{idx}'] = print_razao_social_hp
                                sheet[f'G{idx}'] = mo_valor

                            sheet[f'H{idx}'] = print_ref_nf
                            sheet[f'I{idx}'] = nova_data
                            sheet[f'J{idx}'] = print_valor_vNF
                            sheet[f'K{idx}'] = print_ICMS_prod
                            sheet[f'L{idx}'] = print_valor_vNF * print_ICMS_prod
                            sheet[f'M{idx}'] = print_aliquota_interna
                            sheet[f'N{idx}'] = print_aliquota_interna - print_ICMS_prod

                except AssertionError:
                    pass

                except IndexError:
                    log_instance.log_message(f'Erro ao acessar arquivos: Verifique se todos os arquivos terminam com "-procNFe" ou "-nfe".')
                    continue
        
                except Exception as e:
                    log_instance.log_message(f'Não foi possivel processar arquivos: {e}')

        dados = []
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row = 2, values_only = True):
            dados.append(row)

        dados_sorted = sorted(dados, key=lambda x: x[7])

        for row in sheet.iter_rows(min_row = 2, max_col = sheet.max_column, max_row = sheet.max_row):
            for cell in row:
                cell.value = None

        for col_index, header in enumerate(headers):
            sheet.cell(row = 1, column = col_index + 1, value = header)

        for index, row in enumerate(dados_sorted):
            for col_index, value in enumerate(row):
                sheet.cell(row = index + 2, column = col_index + 1, value = value)
        
        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 10, max_col = 10):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 11, max_col = 11):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00 

        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 12, max_col = 12):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE 

        for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, min_col = 13, max_col = 14):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00     

        for linha in range(2, sheet.max_row + 1):
            for col in range(1, 16):
                celula = sheet.cell(row = linha, column = col)
                celula.fill = PatternFill(start_color = 'FFFFFF', end_color = 'FFFFFF', fill_type = 'solid')
                celula.font = Font(color = '000000', name = 'calibri', size = 11, bold = False)
                celula.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = False)
                celula.border = Border(left=Side(style = 'thin', color = '000000'),
                                    right=Side(style = 'thin', color = '000000'),
                                        top=Side(style = 'thin', color = '000000'),
                                        bottom=Side(style = 'thin', color = '000000'))
                
        workbook.save('Planilha Difal.xlsx')

        if workbook.save:
            try:

                if os.path.exists(f'{download_dir}\\Planilha Difal.xlsx'):
                    os.remove(f'{download_dir}\\Planilha Difal.xlsx')

                shutil.move('Planilha Difal.xlsx', download_dir)

                log_instance.log_message('Planilha criada com sucesso: Planilha Difal')

            except PermissionError:
                log_instance.log_message(f'Erro: permissão para alterar a planilha negada, a planilha já está aberta')
        
            except Exception as e:
                log_instance.log_message(f'Erro: {e}')

        botao_voltar.disabled = False
        botao_voltar.content.color = ft.Colors.WHITE
        page.update()
